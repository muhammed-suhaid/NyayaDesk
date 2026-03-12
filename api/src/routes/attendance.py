from datetime import date, datetime
import io

from flask import Blueprint, request, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from src.db import db
from src.models.advocate import Advocate
from src.models.attendance import Attendance
from src.utils.auth import current_session, require_auth
from src.utils.http import error_response, success_response


attendance_bp = Blueprint("attendance", __name__)


@attendance_bp.post("")
@require_auth
def mark_attendance():
    sess = current_session()
    assert sess is not None
    if sess.role == "super_admin":
        return error_response("Super admin cannot mark attendance", 403)

    if sess.role != "advocate":
        return error_response("Only advocates can mark attendance", 403)

    payload = request.get_json(silent=True) or {}

    from src.models.user import User

    user = User.query.get(sess.user_id)
    if not user:
        return error_response("User not found", 404)

    advocate = Advocate.query.filter_by(email=user.email, company_id=sess.company_id).first()
    if not advocate:
        return error_response("Advocate not found", 404)

    day = date.today()

    record = Attendance.query.filter_by(company_id=sess.company_id, advocate_id=advocate.id, day=day).first()
    if not record:
        record = Attendance(company_id=sess.company_id, advocate_id=advocate.id, day=day)
        db.session.add(record)

    if "checkInTime" in payload:
        t = payload.get("checkInTime")
        record.check_in_time = datetime.strptime(t, "%H:%M").time() if t else None

    if "checkOutTime" in payload:
        t = payload.get("checkOutTime")
        record.check_out_time = datetime.strptime(t, "%H:%M").time() if t else None

    if "status" in payload:
        status = (payload.get("status") or "").strip().lower()
        if status not in {"present", "absent"}:
            return error_response("Invalid attendance status", 400)
        record.status = status

    db.session.commit()
    return success_response("Attendance saved", 201, attendance=record.to_dict())


@attendance_bp.get("")
@require_auth
def list_attendance():
    sess = current_session()
    assert sess is not None
    if sess.role == "super_admin":
        return error_response("Super admin must use /superadmin endpoints", 403)

    advocate_id = request.args.get("advocateId")
    month = request.args.get("month")
    summary = request.args.get("summary")

    q = Attendance.query.filter(Attendance.company_id == sess.company_id)

    if sess.role == "advocate":
        from src.models.user import User

        user = User.query.get(sess.user_id)
        if not user:
            return error_response("User not found", 404)
        advocate = Advocate.query.filter_by(email=user.email, company_id=sess.company_id).first()
        if not advocate:
            return error_response("Advocate not found", 404)
        q = q.filter(Attendance.advocate_id == advocate.id)
    else:
        # admin can optionally filter by advocateId
        if advocate_id:
            q = q.filter(Attendance.advocate_id == int(advocate_id))

    if month:
        try:
            y, m = month.split("-", 1)
            y = int(y)
            m = int(m)
            start = date(y, m, 1)
            end = date(y + (1 if m == 12 else 0), 1 if m == 12 else (m + 1), 1)
            q = q.filter(Attendance.day >= start, Attendance.day < end)
        except Exception:
            return error_response("month must be in YYYY-MM format", 400)

    if summary and sess.role == "admin":
        records = q.all()
        by_adv = {}
        for r in records:
            row = by_adv.setdefault(
                r.advocate_id,
                {
                    "advocateId": r.advocate_id,
                    "advocateName": r.advocate.name if r.advocate else None,
                    "presentDays": 0,
                    "absentDays": 0,
                    "total": 0,
                },
            )
            row["total"] += 1
            if r.status == "present":
                row["presentDays"] += 1
            elif r.status == "absent":
                row["absentDays"] += 1

        return {
            "month": month,
            "summary": list(by_adv.values()),
        }

    records = q.order_by(Attendance.day.desc()).limit(500).all()
    return [r.to_dict() for r in records]


@attendance_bp.get("/export")
@require_auth
def export_attendance():
    sess = current_session()
    assert sess is not None
    if sess.role != "admin":
        return error_response("Only admins can export attendance", 403)

    month_str = request.args.get("month")
    if not month_str:
        month_str = date.today().strftime("%Y-%m")

    try:
        y, m = map(int, month_str.split("-"))
        start = date(y, m, 1)
        # Get next month to find end of current month
        if m == 12:
            end = date(y + 1, 1, 1)
        else:
            end = date(y, m + 1, 1)
        num_days = (end - start).days
    except Exception:
        return error_response("month must be in YYYY-MM format", 400)

    # Fetch advocates for the company
    q_adv = Advocate.query.filter_by(company_id=sess.company_id).order_by(Advocate.name)
    advocate_id = request.args.get("advocateId")
    if advocate_id:
        q_adv = q_adv.filter(Advocate.id == int(advocate_id))
    
    advocates = q_adv.all()
    
    # Fetch all records for the month
    records = Attendance.query.filter(
        Attendance.company_id == sess.company_id,
        Attendance.day >= start,
        Attendance.day < end
    ).all()

    # Matrix: advocate_id -> day -> status
    matrix = {}
    for r in records:
        if r.advocate_id not in matrix:
            matrix[r.advocate_id] = {}
        matrix[r.advocate_id][r.day.day] = "P" if r.status == "present" else "A"

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Define Styles
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    present_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    absent_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
    present_font = Font(color="FFFFFF")
    absent_font = Font(color="FFFFFF")
    
    border = Border(
        left=Side(style='thin', color="D1D5DB"),
        right=Side(style='thin', color="D1D5DB"),
        top=Side(style='thin', color="D1D5DB"),
        bottom=Side(style='thin', color="D1D5DB")
    )

    # Header: Advocate Name, 1, 2, 3, ..., N
    ws.cell(row=1, column=1, value="Advocate Name")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=1).border = border

    for d in range(1, num_days + 1):
        cell = ws.cell(row=1, column=d + 1, value=d)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        # Narrow columns for days
        ws.column_dimensions[cell.column_letter].width = 4

    # Set width for Advocate Name column
    ws.column_dimensions['A'].width = 25

    # Data Rows
    for row_idx, adv in enumerate(advocates, start=2):
        cell_name = ws.cell(row=row_idx, column=1, value=adv.name)
        cell_name.font = Font(bold=True)
        cell_name.border = border
        
        adv_status = matrix.get(adv.id, {})
        for d in range(1, num_days + 1):
            status = adv_status.get(d)
            cell = ws.cell(row=row_idx, column=d + 1, value=status if status else "")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            
            if status == "P":
                cell.fill = present_fill
                cell.font = present_font
            elif status == "A":
                cell.fill = absent_fill
                cell.font = absent_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Filename like: Attendance_March_2026.xlsx
    formatted_month = start.strftime("%B_%Y")
    filename = f"Attendance_Report_{formatted_month}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={filename}"},
    )
